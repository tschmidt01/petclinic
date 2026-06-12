from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment

# --- Data pulled from PostgreSQL via MCP ---
owners = [
    (11, "George", "Darling", "London", 2),
    (8, "Alice", "Liddell", "Oxford", 2),
    (15, "Beatrix", "Potter", "Near Sawrey", 2),
    (6, "Roger", "Radcliff", "London", 2),
    (24, "Lady", "Tremaine", "Ile-de-France", 2),
    (18, "Wallace", "Wensleydale", "Wigan", 2),
    (9, "Henry", "Baskerville", "Dartmoor", 1),
    (14, "Sam", "Carraclough", "Yorkshire", 1),
    (19, "Wendy", "Darling", "London", 1),
    (27, "Charles", "Dickens", "Higham", 1),
    (10, "John", "Dolittle", "Puddleby", 1),
    (17, "Argus", "Filch", "Inverness", 1),
    (25, "Mister", "Geppetto", "Florence", 1),
    (21, "Hermione", "Granger", "Hogsmeade", 1),
    (20, "Rubeus", "Hagrid", "Hogsmeade", 1),
    (28, "Sherlock", "Holmes", "London", 1),
    (1, "Kevin", "McCallister", "Winnetka", 1),
    (2, "Harry", "Potter", "Little Whinging", 1),
    (26, "Alonso", "Quixano", "La Mancha", 1),
    (23, "Tintin", "Reporter", "Brussels", 1),
    (4, "Tom", "Riddle", "Wiltshire", 1),
    (7, "Newt", "Scamander", "London", 1),
    (3, "Erwin", "Schroedinger", "Vienna", 1),
    (16, "Long", "Silver", "Bristol", 1),
    (22, "Salazar", "Slytherin", "Hogsmeade", 1),
    (5, "Ronald", "Weasley", "Ottery St Catchpole", 1),
    (12, "James", "Bond", "London", 0),
    (13, "Hercule", "Poirot", "London", 0),
]

pet_types = [
    ("dog", 10),
    ("cat", 9),
    ("hamster", 4),
    ("bird", 4),
    ("snake", 2),
    ("lizard", 2),
    ("horse", 1),
]

HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(bold=True, color="FFFFFF")

wb = Workbook()

# --- Sheet 1: Owners by pet count ---
ws = wb.active
ws.title = "Owners by Pets"
headers = ["Owner ID", "First Name", "Last Name", "City", "Pet Count"]
ws.append(headers)
for c in ws[1]:
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal="center")

for row in owners:
    ws.append(list(row))

widths = [10, 14, 16, 22, 11]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[chr(64 + i)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:E{len(owners) + 1}"

# --- Sheet 2: Pet types + pie chart ---
ws2 = wb.create_sheet("Pet Types")
ws2.append(["Pet Type", "Count"])
for c in ws2[1]:
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal="center")
for row in pet_types:
    ws2.append(list(row))
ws2.column_dimensions["A"].width = 14
ws2.column_dimensions["B"].width = 10

total = sum(c for _, c in pet_types)
ws2.append([])
ws2.append(["Total", total])
ws2[f"A{len(pet_types) + 3}"].font = Font(bold=True)
ws2[f"B{len(pet_types) + 3}"].font = Font(bold=True)

pie = PieChart()
labels = Reference(ws2, min_col=1, min_row=2, max_row=len(pet_types) + 1)
data = Reference(ws2, min_col=2, min_row=1, max_row=len(pet_types) + 1)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.title = "Proportion of Pet Types"
pie.height = 10
pie.width = 14
pie.dataLabels = DataLabelList()
pie.dataLabels.showPercent = True
ws2.add_chart(pie, "D2")

out = "owners-pets-report.xlsx"
wb.save(out)
print(f"Wrote {out}: {len(owners)} owners, {len(pet_types)} pet types, total {total} pets")
